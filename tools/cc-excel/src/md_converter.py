"""Convert Excel (XLSX) files to Markdown tables."""

from pathlib import Path

import openpyxl


def convert_xlsx_to_markdown(
    input_path: Path,
    sheet_name: str | None = None,
    all_sheets: bool = False,
) -> str:
    """Convert an Excel workbook to Markdown with pipe tables.

    Args:
        input_path: Path to the ``.xlsx`` file.
        sheet_name: Name of a specific sheet to convert.  When *None* and
            *all_sheets* is *False*, only the first sheet is converted.
        all_sheets: If *True*, convert every sheet in the workbook.

    Returns:
        Markdown string with one ``## Sheet Name`` heading and pipe table
        per sheet.
    """
    wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {available}"
            )
        sheets = [wb[sheet_name]]
    elif all_sheets:
        sheets = [wb[name] for name in wb.sheetnames]
    else:
        sheets = [wb[wb.sheetnames[0]]]

    sections: list[str] = []

    for ws in sheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sections.append(f"## {ws.title}\n\n*Empty sheet*\n")
            continue

        # First row is header
        headers = [_cell_to_str(c) for c in rows[0]]
        col_count = len(headers)

        # Calculate column widths for alignment
        widths = [len(h) for h in headers]
        data_rows: list[list[str]] = []
        for row in rows[1:]:
            cells = [_cell_to_str(c) for c in row[:col_count]]
            # Pad if row has fewer columns
            while len(cells) < col_count:
                cells.append("")
            data_rows.append(cells)
            for i, cell in enumerate(cells):
                widths[i] = max(widths[i], len(cell))

        # Ensure minimum width of 3 for separator
        widths = [max(w, 3) for w in widths]

        # Build table
        lines: list[str] = []
        # Header
        header_line = "| " + " | ".join(h.ljust(widths[i]) for i, h in enumerate(headers)) + " |"
        lines.append(header_line)
        # Separator
        sep_line = "| " + " | ".join("-" * widths[i] for i in range(col_count)) + " |"
        lines.append(sep_line)
        # Data rows
        for row_cells in data_rows:
            row_line = "| " + " | ".join(row_cells[i].ljust(widths[i]) for i in range(col_count)) + " |"
            lines.append(row_line)

        table_md = "\n".join(lines)
        sections.append(f"## {ws.title}\n\n{table_md}\n")

    wb.close()
    return "\n".join(sections).strip() + "\n"


def _cell_to_str(value: object) -> str:
    """Convert a cell value to a clean string representation."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Remove trailing zeros for cleaner output
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).replace("|", "\\|").replace("\n", " ")
